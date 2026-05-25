from __future__ import annotations

import csv
import io
import sys
import zipfile
from typing import Any, Dict, Tuple
from xml.etree import ElementTree


MAX_AGENT_SOURCE_FILE_BYTES = 10 * 1024 * 1024
MAX_EXTRACTED_TEXT_CHARS = 30000
SUPPORTED_AGENT_SOURCE_EXTENSIONS_HINT = "TXT, CSV, TSV, MD, PDF, DOCX, XLSX"
SUPPORTED_AGENT_SOURCE_EXTENSIONS = {".txt", ".md", ".csv", ".tsv", ".pdf", ".docx", ".xlsx"}
SUPPORTED_AGENT_SOURCE_MIME_PREFIXES = {
    "text/",
}
SUPPORTED_AGENT_SOURCE_MIME_TYPES = {
    "",
    "application/octet-stream",
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "text/csv",
    "text/tab-separated-values",
    "text/markdown",
    "text/plain",
}


def build_agent_source_from_upload(file_storage: Any, preferred_name: str = "") -> Tuple[Dict[str, Any], Dict[str, Any]]:
    if not file_storage:
        return {}, {"code": "FILE_REQUIRED", "message": "Выберите файл для агента."}
    file_name = _clean_text(getattr(file_storage, "filename", "")) or "uploaded-file"
    extension = _file_extension(file_name)
    if extension not in SUPPORTED_AGENT_SOURCE_EXTENSIONS:
        return {}, {
            "code": "UNSUPPORTED_FILE_TYPE",
            "message": f"Этот тип файла пока не поддерживается. Загрузите {SUPPORTED_AGENT_SOURCE_EXTENSIONS_HINT}.",
            "supported_extensions": sorted(SUPPORTED_AGENT_SOURCE_EXTENSIONS),
        }
    mime_type = _clean_text(getattr(file_storage, "mimetype", ""))
    if not _is_supported_mime(mime_type):
        return {}, {
            "code": "UNSUPPORTED_MIME_TYPE",
            "message": f"Формат файла не похож на поддерживаемый документ. Поддерживаются {SUPPORTED_AGENT_SOURCE_EXTENSIONS_HINT}.",
            "supported_extensions": sorted(SUPPORTED_AGENT_SOURCE_EXTENSIONS),
        }

    data = file_storage.read()
    if len(data) > MAX_AGENT_SOURCE_FILE_BYTES:
        return {}, {
            "code": "FILE_TOO_LARGE",
            "message": "Файл слишком большой. Максимальный размер для агента - 10 МБ.",
            "max_bytes": MAX_AGENT_SOURCE_FILE_BYTES,
        }
    if not data:
        return {}, {"code": "EMPTY_FILE", "message": "Файл пустой. Добавьте документ с текстом или вставьте текст вручную."}

    extracted_text, extraction_error = extract_text_from_agent_source_bytes(data, file_name, mime_type)
    if extraction_error:
        return {}, extraction_error
    if not _clean_text(extracted_text):
        return {}, {
            "code": "EMPTY_EXTRACTION",
            "message": "Файл загружен, но текст извлечь не удалось. Попробуйте DOCX/TXT или вставьте текст вручную.",
        }

    return {
        "source_type": "file",
        "name": _clean_text(preferred_name) or file_name,
        "file_name": file_name,
        "mime_type": mime_type,
        "file_size_bytes": len(data),
        "content_text": extracted_text[:MAX_EXTRACTED_TEXT_CHARS],
        "extraction_state": "ready",
        "extraction_method": _extraction_method(extension),
    }, {}


def extract_text_from_agent_source_bytes(data: bytes, file_name: str, mime_type: str = "") -> Tuple[str, Dict[str, Any]]:
    extension = _file_extension(file_name)
    try:
        if extension in {".txt", ".md", ".csv", ".tsv"}:
            return _decode_text(data), {}
        if extension == ".pdf":
            return _extract_pdf_text(data), {}
        if extension == ".docx":
            return _extract_docx_text(data), {}
        if extension == ".xlsx":
            return _extract_xlsx_text(data), {}
    except Exception:
        exc = sys.exc_info()[1]
        return "", {
            "code": "EXTRACTION_FAILED",
            "message": "Не удалось извлечь текст из файла. Попробуйте другой формат или вставьте текст вручную.",
            "details": str(exc)[:200],
            "mime_type": mime_type,
        }
    return "", {
        "code": "UNSUPPORTED_FILE_TYPE",
        "message": f"Этот тип файла пока не поддерживается. Загрузите {SUPPORTED_AGENT_SOURCE_EXTENSIONS_HINT}.",
    }


def _extract_pdf_text(data: bytes) -> str:
    try:
        from pypdf import PdfReader
    except Exception:
        raise RuntimeError("pypdf is required for PDF extraction")
    reader = PdfReader(io.BytesIO(data))
    pages = []
    for index, page in enumerate(reader.pages):
        if index >= 50:
            break
        text = page.extract_text() or ""
        if text.strip():
            pages.append(text.strip())
    return "\n\n".join(pages)


def _extract_docx_text(data: bytes) -> str:
    archive = zipfile.ZipFile(io.BytesIO(data))
    try:
        xml_data = archive.read("word/document.xml")
    finally:
        archive.close()
    root = ElementTree.fromstring(xml_data)
    namespace = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    lines = []
    for paragraph in root.iter(f"{namespace}p"):
        fragments = []
        for text_node in paragraph.iter(f"{namespace}t"):
            if text_node.text:
                fragments.append(text_node.text)
        line = "".join(fragments).strip()
        if line:
            lines.append(line)
    return "\n".join(lines)


def _extract_xlsx_text(data: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except Exception:
        raise RuntimeError("openpyxl is required for XLSX extraction")
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    output = io.StringIO()
    writer = csv.writer(output)
    for sheet_index, sheet in enumerate(workbook.worksheets):
        if sheet_index >= 5:
            break
        writer.writerow([f"sheet: {sheet.title}"])
        for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_index >= 200:
                break
            values = [_clean_text(cell) for cell in row]
            if any(values):
                writer.writerow(values)
    return output.getvalue()


def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _is_supported_mime(mime_type: str) -> bool:
    if mime_type in SUPPORTED_AGENT_SOURCE_MIME_TYPES:
        return True
    return any(mime_type.startswith(prefix) for prefix in SUPPORTED_AGENT_SOURCE_MIME_PREFIXES)


def _extraction_method(extension: str) -> str:
    return {
        ".txt": "plain_text",
        ".md": "plain_text",
        ".csv": "plain_text_csv",
        ".tsv": "plain_text_tsv",
        ".pdf": "pypdf",
        ".docx": "docx_xml",
        ".xlsx": "openpyxl",
    }.get(extension, "unknown")


def _file_extension(file_name: str) -> str:
    if "." not in file_name:
        return ""
    return "." + file_name.rsplit(".", 1)[-1].lower()


def _clean_text(value: Any) -> str:
    return str(value or "").strip()
